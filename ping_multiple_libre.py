import socket
from ping3 import ping
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

# Function to resolve hostname to IP
def resolve_hostname(hostname):
    try:
        ip_address = socket.gethostbyname(hostname)
        return ip_address
    except socket.gaierror as e:
        print(f"Error resolving {hostname}: {e}")
        return None

# Function to color code cells based on ping results
def get_color_for_latency(latency):
    if latency is None or latency > 0.150:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for latency above 150ms or no response
    elif latency > 0.050:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for latency between 50ms and 150ms
    else:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for latency below 50ms

# Function to ping the host and get results
def ping_target(ip_address, count):
    latencies = []
    for _ in range(count):
        latency = ping(ip_address)
        latencies.append(latency)
    return latencies

# Function to create an Excel spreadsheet with the results
def create_spreadsheet(results, output_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ping Results"

    # Write headers
    headers = ["Host/IP", "Latency (ms)", "Status"]
    ws.append(headers)

    # Write ping results
    for result in results:
        host_ip, latencies = result
        average_latency = sum([latency for latency in latencies if latency is not None]) / len(latencies)
        status = "OK" if average_latency <= 0.150 else "Failed"

        # Write the host/IP, average latency, and status
        row = [host_ip, f"{average_latency:.4f}", status]
        ws.append(row)

        # Apply color to the latency cell based on the value
        color = get_color_for_latency(average_latency)
        latency_cell = ws.cell(row=ws.max_row, column=2)
        latency_cell.fill = color

    # Save the workbook
    wb.save(output_filename)
    print(f"Results saved to {output_filename}")

# Main function
def main():
    # Step 1: Get the input file from the user
    input_file = input("Enter the name of the text file with the list of hostnames/IPs (e.g., hosts.txt): ").strip()

    try:
        with open(input_file, "r") as file:
            hosts = file.readlines()
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        return

    # Step 2: Get the number of pings to send
    try:
        count = int(input("How many ping requests would you like to send to each host? ").strip())
        if count < 1:
            raise ValueError("The number of ping requests must be at least 1.")
    except ValueError as e:
        print(f"Invalid input: {e}")
        return

    # Step 3: Collect the ping results
    results = []
    for host in hosts:
        host = host.strip()  # Remove extra whitespace/newlines
        if not host.replace('.', '').isdigit():  # If it's a hostname
            ip_address = resolve_hostname(host)
            if ip_address:
                latencies = ping_target(ip_address, count)
                results.append((host, latencies))
        else:  # If it's an IP address
            latencies = ping_target(host, count)
            results.append((host, latencies))

    # Step 4: Create the Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"ping_results_{timestamp}.xlsx"
    create_spreadsheet(results, output_filename)

if __name__ == "__main__":
    main()