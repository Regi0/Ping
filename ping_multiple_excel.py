import socket
from ping3 import ping
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import os

# Function to resolve hostname to IP
def resolve_hostname(hostname):
    try:
        ip_address = socket.gethostbyname(hostname)
        return ip_address
    except socket.gaierror as e:
        print(f"Error resolving {hostname}: {e}")
        return None

# Function to color code cells based on ping results
def get_color_for_latency(latency):
    if latency is None or latency > 0.150:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for latency above 150ms or no response
    elif latency > 0.050:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for latency between 50ms and 150ms
    else:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for latency below 50ms

# Function to ping the host and get results
def ping_target(ip_address, count):
    latencies = []
    for _ in range(count):
        latency = ping(ip_address)
        latencies.append(latency)
    return latencies

# Function to create an Excel spreadsheet with the results
def create_spreadsheet(results, output_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ping Results"

    # Write headers
    headers = ["Host/IP", "Resolved IP", "Pings Sent", "Average Latency (ms)", "Status"]
    ws.append(headers)

    # Write ping results
    for result in results:
        host_ip, latencies, resolved_ip, ping_count = result
        
        # Remove None values for latencies that failed
        valid_latencies = [latency for latency in latencies if latency is not None]
        
        if valid_latencies:
            average_latency = sum(valid_latencies) / len(valid_latencies)
        else:
            average_latency = None
        
        status = "OK" if average_latency is not None and average_latency <= 0.150 else "Failed"

        # Write the host/IP, resolved IP, number of pings sent, average latency, and status
        row = [host_ip, resolved_ip if resolved_ip else "N/A", ping_count, f"{average_latency:.4f}" if average_latency is not None else "No Response", status]
        ws.append(row)

        # Apply color to the average latency cell based on the value
        color = get_color_for_latency(average_latency)
        latency_cell = ws.cell(row=ws.max_row, column=4)  # Column for average latency
        latency_cell.fill = color

    # Save the workbook
    wb.save(output_filename)
    print(f"Results saved to {output_filename}")

# Main function
def main():
    # Step 1: Get the input Excel file from the user
    input_file = input("Enter the name of the Excel file with the list of hostnames/IPs (e.g., hosts.xlsx): ").strip()

    try:
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        hosts = [cell.value for cell in sheet['A'] if cell.value]  # Assuming the first column contains the hostnames/IPs
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        return
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return

    # Step 2: Get the number of pings to send
    try:
        count = int(input("How many ping requests would you like to send to each host? ").strip())
        if count < 1:
            raise ValueError("The number of ping requests must be at least 1.")
    except ValueError as e:
        print(f"Invalid input: {e}")
        return

    # Step 3: Collect the ping results
    results = []
    for host in hosts:
        host = host.strip()  # Remove extra whitespace/newlines
        resolved_ip = None
        if not host.replace('.', '').isdigit():  # If it's a hostname
            resolved_ip = resolve_hostname(host)
            if resolved_ip:
                latencies = ping_target(resolved_ip, count)
                results.append((host, latencies, resolved_ip, count))
        else:  # If it's an IP address
            latencies = ping_target(host, count)
            results.append((host, latencies, host, count))

    # Step 4: Create the Excel file with the new naming convention
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = os.path.splitext(os.path.basename(input_file))[0]  # Get the name of the input file without extension
    output_filename = f"ping_results_{base_filename}_{timestamp}.xlsx"
    
    # Create the spreadsheet and save it
    create_spreadsheet(results, output_filename)

if __name__ == "__main__":
    main()